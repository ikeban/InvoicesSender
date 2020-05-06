import openpyxl

class ExcelReader:
    """ Reads all needed data from excel file """
    def __init__(self, excelFileName='InvoiceSenderControl.xlsx', sheet=None):
        self._sheet = sheet or openpyxl.load_workbook(excelFileName)['Control']
        self._data = []
        self._smtpData = self._parseSmtp()
        self._keyWords = self._parseKeyWords()
        self._readAllLines()

    def getData(self):
        return self._data

    def getSmtpData(self):
        return self._smtpData

    def _parseSmtp(self):
        smtpAddress = self._sheet.cell(row = 2, column = 4).value
        smtpPort = self._sheet.cell(row = 2, column = 5).value
        ownerAddress = self._sheet.cell(row = 2, column = 6).value
        ownerPassword = self._sheet.cell(row = 2, column = 7).value
        return (smtpAddress, smtpPort, ownerAddress, ownerPassword)

    def _parseKeyWords(self):
        columnIndex = 8
        keyWords = []
        while True:
            keyWord = self._sheet.cell(row = 4, column = columnIndex).value
            if self._checkIfKeyWord(keyWord) == False:
                return keyWords
            keyWords.append(keyWord)
            columnIndex += 1
            

    def _checkIfKeyWord(self, keyWord):
        return (keyWord is not None) and len(keyWord) >= 3 and keyWord[0] == '[' and keyWord[-1] == ']'

    def _readAllLines(self):
        rowIndex = 5
        while True:
            invoiceText = self._sheet.cell(row = rowIndex, column = 2).value
            if not invoiceText:
                return
            self._readOneLine(rowIndex)
            rowIndex += 1

    def _readOneLine(self, rowNumber):
        invoiceText = self._sheet.cell(row = rowNumber, column = 2).value
        emailAddress = self._sheet.cell(row = rowNumber, column = 3).value
        templateName = self._sheet.cell(row = rowNumber, column = 4).value
        emailSubject = self._sheet.cell(row = rowNumber, column = 5).value
        messageId = self._sheet.cell(row = rowNumber, column = 6).value

        keyWordMap = {}
        for index in range(len(self._keyWords)):
            keyWordValue = self._sheet.cell(row = rowNumber, column = 8 + index).value
            keyWordMap[ self._keyWords[index] ] = keyWordValue

        self._data.append( (invoiceText, emailAddress, templateName, keyWordMap, emailSubject, messageId) )
                    
